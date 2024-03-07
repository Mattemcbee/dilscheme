import openpyxl
import datetime
import os

def filter_patients(file1_path):
    wb1 = openpyxl.load_workbook(file1_path)
    sheet1 = wb1.active

    drugName = 'drug'
    timePointColumn = 'Timepoint'
    DilutionColumn = 'Dilution Factor'
    subject_col_name = 'Subject ID'

    Timepoint_index = None
    for col_index, cell in enumerate(sheet1[1], start=1):
        if cell.value == timePointColumn:
            Timepoint_index = col_index
            print('column ' + str(Timepoint_index))
            break

    if Timepoint_index is None:
        print(f"Timepoint column not found ({timePointColumn}) in the first file.")

    Dilution_index = None
    for col_index, cell in enumerate(sheet1[1], start=1):
        if cell.value == DilutionColumn:
            Dilution_index = col_index
            print('column ' + str(Dilution_index))
            break

    if Dilution_index is None:
        print(f"Dilution column not found ({DilutionColumn}) in the first file.")

    allTimepoints = {}

    for row in sheet1.iter_rows(min_row=2, values_only=True):
        timepoint = row[Timepoint_index - 1]
        dilution = row[Dilution_index - 1]

        # Check if the timepoint is "PRE"
        if timepoint == "PRE":
          allTimepoints[timepoint] = 1  # Set dilution to 1 for "PRE"
        elif timepoint == None:
          allTimepoints[timepoint] = ""  # Set dilution to blank for blank timepoint
        elif timepoint == "":
          allTimepoints[timepoint] = ""  # Set
        elif dilution is not None and (timepoint not in allTimepoints or dilution > allTimepoints[timepoint]):
          allTimepoints[timepoint] = dilution

    print(allTimepoints)

    # Iterate through the rows to fill empty cells in the dilution column
    for row in sheet1.iter_rows(min_row=2, values_only=False):
        timepoint = row[Timepoint_index - 1].value
        dilution_cell = row[Dilution_index - 1]

        # Check if the dilution cell is empty
        if dilution_cell.value is None:
            # Fill the dilution cell with the highest dilution for the corresponding timepoint
            dilution_cell.value = allTimepoints.get(timepoint, None)

    # Create a new filename for the modified file
    modified_file_path = "modified_" + os.path.basename(file1_path)

    # Save the modified workbook with the new filename
    wb1.save(modified_file_path)
    print(f"Modified file saved to {modified_file_path}")

# Example usage
filter_patients("dil_template.xlsx")
